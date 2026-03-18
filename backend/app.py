from flask import Flask, render_template, request, redirect, session, url_for, g, flash, send_file
from flask_bcrypt import Bcrypt
from flask_cors import CORS
from db import get_db_connection
import pandas as pd
import qrcode
import io
import base64
from datetime import datetime




app = Flask(__name__)
app.secret_key = "secret123"

bcrypt = Bcrypt(app)
CORS(app)

def get_db():
    """Get a database connection for the current request."""
    if 'db' not in g:
        try:
            g.db = get_db_connection()
        except Exception as e:
            print(f"Database connection failed: {e}")
            flash("Database connection error. Please try again later.", "error")
            raise
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    """Close the database connection at the end of the request."""
    db = g.pop('db', None)
    if db is not None:
        try:
            db.close()
        except Exception as close_error:
            print(f"Error closing database: {close_error}")
@app.route("/create-test-user")
def create_test_user():
    db = get_db()
    cur = db.cursor()

    hashed = bcrypt.generate_password_hash("admin123").decode('utf-8')

    cur.execute("""
        INSERT INTO tbl_user (User_name, Email, password_hash, user_type)
        VALUES (%s, %s, %s, %s)
    """, ("admin", "admin@gmail.com", hashed, "Admin"))  # Changed to use bcrypt hash

    db.commit()

    return "Test user created"
@app.route("/")
def home():
    return render_template("home.html")

@app.route("/free-qr")
def free_qr():
    return render_template("free_qr.html")


@app.route("/set-admin-password")
def set_admin_password():
    db = get_db()
    cur = db.cursor()

    hashed = bcrypt.generate_password_hash("admin123").decode('utf-8')

    cur.execute("""
        UPDATE tbl_user
        SET password_hash = %s
        WHERE Email = %s
    """, (hashed, "admin@gmail.com"))  # Changed to use bcrypt hash

    db.commit()

    return "Admin password set"


# ---------------- REGISTER ----------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["full_name"]
        email = request.form["email"]
        password = request.form["password"]
        user_type = request.form["user_type"]

        hashed_pw = bcrypt.generate_password_hash(password).decode("utf-8")

        db = get_db()
        cursor = db.cursor()

        cursor.execute(
            "INSERT INTO tbl_user (User_name, Email, password_hash, user_type) VALUES (%s, %s, %s, %s)",
            (username, email, hashed_pw, user_type)
        )
        db.commit()

        return redirect("/login")

    return render_template("register.html")


# ---------------- LOGIN ----------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        db = get_db()
        cursor = db.cursor(dictionary=True, buffered=True)

        cursor.execute("SELECT * FROM tbl_user WHERE Email=%s", (email,))
        user = cursor.fetchone()

        if user and bcrypt.check_password_hash(user["password_hash"], password):
            session["user_id"] = user["User_id"]
            session["username"] = user["User_name"]
            session["role"] = user["user_type"]
            return redirect("/dashboard")

        flash("Invalid email or password. Please try again.", "error")
        return redirect("/login")

    return render_template("login.html")


# ---------------- DASHBOARD ----------------
@app.route("/dashboard")
def dashboard():
    if "user_id" not in session:
        return redirect("/login")

    if session.get("role") == "Admin":
        return redirect("/admin/dashboard")

    # User Dashboard Logic
    db = get_db()
    cur = db.cursor(dictionary=True, buffered=True)
    
    # Fetch categories for the dropdown
    cur.execute("SELECT * FROM tbl_category")
    categories = cur.fetchall()

    # Fetch assets belonging to the logged-in user
    cur.execute("SELECT a.*, c.Category_name FROM tbl_asset a LEFT JOIN tbl_category c ON a.Category_ID = c.Category_id WHERE a.user_id = %s ORDER BY a.Sl_no DESC", (session["user_id"],))
    assets = cur.fetchall()

    # Fetch expired assets for the logged-in user
    cur.execute("SELECT * FROM tbl_asset WHERE expiry_date < CURDATE() AND user_id = %s", (session["user_id"],))
    expired_assets = cur.fetchall()

    return render_template(
        "user_dashboard.html",
        username=session["username"],
        categories=categories,
        assets=assets,
        expired_assets=expired_assets
    )


@app.route("/admin/dashboard")
def admin_dashboard():
    if "user_id" not in session or session["role"] != "Admin":
        return redirect("/login")

    db = get_db()
    cur = db.cursor(dictionary=True, buffered=True)

    # 1. KPIs
    cur.execute("SELECT COUNT(*) as count FROM tbl_asset")
    total_assets = cur.fetchone()['count'] if cur.rowcount > 0 else 0

    cur.execute("SELECT COUNT(*) as count FROM tbl_category")
    total_categories = cur.fetchone()['count'] if cur.rowcount > 0 else 0

    cur.execute("SELECT COUNT(*) as count FROM tbl_feedback")
    pending_requests = cur.fetchone()['count'] if cur.rowcount > 0 else 0

    cur.execute("SELECT COUNT(*) as count FROM tbl_asset WHERE expiry_date < CURDATE()")
    expired_assets_count = cur.fetchone()['count'] if cur.rowcount > 0 else 0

    # 2. Categories with Asset Counts
    cur.execute("""
        SELECT c.Category_id, c.Category_name, COUNT(a.Sl_no) as asset_count
        FROM tbl_category c
        LEFT JOIN tbl_asset a ON c.Category_id = a.Category_ID
        GROUP BY c.Category_id
    """)
    categories = cur.fetchall()

    # 3. Recent Feedbacks
    cur.execute("""
        SELECT f.*, u.User_name
        FROM tbl_feedback f
        JOIN tbl_user u ON f.user_id = u.User_id
        ORDER BY f.feedback_id DESC LIMIT 5
    """)
    feedbacks = cur.fetchall()

    # 4. Expired Assets
    cur.execute("""
        SELECT * FROM tbl_asset
        WHERE expiry_date < CURDATE()
        LIMIT 5
    """)
    expired_assets = cur.fetchall()

    # 5. All Assets for QR Report
    cur.execute("""
        SELECT a.*, c.Category_name 
        FROM tbl_asset a 
        LEFT JOIN tbl_category c ON a.Category_ID = c.Category_id 
        ORDER BY a.Sl_no DESC
    """)
    all_assets = cur.fetchall()

    return render_template(
        "dashboard.html",
        username=session["username"],
        total_assets=total_assets,
        total_categories=total_categories,
        pending_requests=pending_requests,
        expired_assets_count=expired_assets_count,
        categories=categories,
        feedbacks=feedbacks,
        expired_assets=expired_assets,
        all_assets=all_assets
    )

# ---------------- ADMIN ACTIONS ----------------

@app.route("/admin/add-category", methods=["POST"])
def add_category():
    if "user_id" not in session or session["role"] != "Admin": return redirect("/login")
    category_name = request.form.get("category_name")
    if category_name:
        db = get_db()
        cur = db.cursor()
        cur.execute("INSERT INTO tbl_category (Category_name) VALUES (%s)", (category_name,))
        db.commit()
        flash("Category added successfully!", "success")
    return redirect("/admin/dashboard")

@app.route("/admin/delete-category/<int:id>")
def delete_category(id):
    if "user_id" not in session or session["role"] != "Admin": return redirect("/login")
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM tbl_category WHERE Category_id = %s", (id,))
    db.commit()
    flash("Category deleted.", "success")
    return redirect("/admin/dashboard")

@app.route("/add-asset", methods=["POST"])
def add_asset():
    if "user_id" not in session: return redirect("/login")
    # Manual Add
    name = request.form.get("asset_name")
    cat_id = request.form.get("category_id")
    expiry = request.form.get("expiry_date")
    
    # Additional Fields
    manufacturer = request.form.get("manufacturer")
    vendor = request.form.get("vendor")
    price = request.form.get("price")
    bill_no = request.form.get("bill_no")
    purchase_date = request.form.get("purchase_date")
    custodian = request.form.get("custodian")
    department = request.form.get("department")
    location = request.form.get("location")
    warranty_period = request.form.get("warranty_period")
    stock_register_no = request.form.get("stock_register_no")
    description = request.form.get("description")
    
    # Generate QR Data (Simple string for now)
    qr_data = f"ASSET-{name}-{datetime.now().timestamp()}"
    # Generate a Product ID if needed, or use random
    prod_id = f"PID-{int(datetime.now().timestamp())}"
    
    db = get_db()
    cur = db.cursor()
    cur.execute("""
        INSERT INTO tbl_asset (
            Product_ID, Product_name, Category_ID, Manufacturer, Vendor, Price, 
            Bill_no, Purchase_date, expiry_date, custodian, Department, Location, 
            Warranty_period, Stock_register_no, Description, 
            Qr_code_value, Status, created_by, user_id
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Active', %s, %s)
    """, (
        prod_id, name, cat_id, manufacturer, vendor, price,
        bill_no, purchase_date, expiry, custodian, department, location,
        warranty_period, stock_register_no, description,
        qr_data, session['user_id'], session['user_id']
    ))
    db.commit()
    flash("Asset added successfully!", "success")
    return redirect("/dashboard")

@app.route("/upload-bulk", methods=["POST"])
def upload_bulk():
    if "user_id" not in session: return redirect("/login")
    file = request.files.get('file')
    if file and file.filename.endswith(('.xls', '.xlsx')):
        try:
            df = pd.read_excel(file)
            # Replace NaN values with None (NULL in SQL)
            df = df.where(pd.notnull(df), None)
            
            db = get_db()
            cur = db.cursor()
            for _, row in df.iterrows():
                # Generate QR Data
                qr_data = f"ASSET-{row.get('Product_name')}-{datetime.now().timestamp()}"
                
                # Use provided Product_ID or generate one if missing
                prod_id = row.get('Product_ID')
                if not prod_id:
                    prod_id = f"PID-{int(datetime.now().timestamp())}"

                cur.execute("""
                    INSERT INTO tbl_asset (
                        Product_ID, Product_name, Category_ID, Manufacturer, Vendor, Price, 
                        Bill_no, Purchase_date, expiry_date, custodian, Department, Location, 
                        Warranty_period, Stock_register_no, Description, 
                        Qr_code_value, Status, created_by, user_id
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Active', %s, %s)
                """, (
                    prod_id, row.get('Product_name'), row.get('Category_ID'), row.get('Manufacturer'),
                    row.get('Vendor'), row.get('Price'), row.get('Bill_no'), row.get('Purchase_date'),
                    row.get('expiry_date'), row.get('custodian'), row.get('Department'), row.get('Location'),
                    row.get('Warranty_period'), row.get('Stock_register_no'), row.get('Description'),
                    qr_data, session['user_id'], session['user_id']
                ))
            db.commit()
            flash("Bulk upload successful!", "success")
        except Exception as e:
            flash(f"Error uploading file: {str(e)}", "error")
    return redirect("/dashboard")

@app.route("/submit-feedback", methods=["POST"])
def submit_feedback():
    if "user_id" not in session: return redirect("/login")
    
    message = request.form.get("message")
    
    if message:
        db = get_db()
        cur = db.cursor()
        cur.execute("INSERT INTO tbl_feedback (user_id, message) VALUES (%s, %s)",
                    (session["user_id"], message))
        db.commit()
        flash("Feedback submitted successfully.", "success")
    return redirect("/dashboard")

@app.route("/admin/delete-feedback/<int:id>")
def delete_feedback(id):
    if "user_id" not in session or session["role"] != "Admin": return redirect("/login")
    db = get_db()
    cur = db.cursor()
    cur.execute("DELETE FROM tbl_feedback WHERE feedback_id=%s", (id,))
    db.commit()
    flash("Feedback deleted.", "success")
    return redirect("/admin/dashboard")

@app.route("/admin/export-report")
def export_report():
    if "user_id" not in session or session["role"] != "Admin": return redirect("/login")
    db = get_db()
    df = pd.read_sql("SELECT * FROM tbl_asset", db)
    # Convert date columns to proper string format (YYYY-MM-DD)
    if 'Purchase_date' in df.columns:
        df['Purchase_date'] = pd.to_datetime(df['Purchase_date'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')
    if 'expiry_date' in df.columns:
        df['expiry_date'] = pd.to_datetime(df['expiry_date'], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')
    output = io.BytesIO()
    # Export as Excel with proper column widths
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Assets')
        worksheet = writer.sheets['Assets']
        # Auto-adjust column widths
        for idx, col in enumerate(df.columns):
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 30)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="asset_report.xlsx")

@app.route("/admin/export-user-report")
def export_user_report():
    if "user_id" not in session or session["role"] != "Admin": return redirect("/login")
    db = get_db()
    df = pd.read_sql("SELECT User_id, User_name, Email, user_type, Phone_no FROM tbl_user", db)
    output = io.BytesIO()
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Users')
        worksheet = writer.sheets['Users']
        for idx, col in enumerate(df.columns):
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 30)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="user_report.xlsx")

# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


if __name__ == "__main__":
    app.run(debug=True)
